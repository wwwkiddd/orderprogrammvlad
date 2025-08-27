
# -*- coding: utf-8 -*-

"""
Наряд-Заказ — v2.3
Исправления и улучшения:
- Админ-операции (добавить компанию/номер, выставить оплату, удалить) работают даже когда окно «Создать наряд» НЕ открыто.
- Больше нет ошибок invalid command name при обновлении списков.
- Тумблер «Оплата» в админке корректно отражает состояние из файла и сразу обновляется при выборе компании/поиске.
- Компании в списках идут в том же порядке, что и в файле; новые добавляются В КОНЕЦ.
"""

import os
import datetime
import subprocess
from pathlib import Path
import tkinter as tk
from tkinter import BOTH, LEFT, RIGHT, Y, X, NW, DISABLED, NORMAL, messagebox, simpledialog
from tkinter import ttk

from openpyxl import load_workbook
from num2words import num2words
import pandas as pd
import ttkbootstrap as tb

# === Пути проекта ===
BASE_DIR = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / "templates"
OUTPUT_DIR = BASE_DIR / "output"
DATA_DIR = BASE_DIR / "data"
TEMPLATE_XLSX = TEMPLATES_DIR / "order_template.xlsx"
COMPANIES_XLSX = DATA_DIR / "companies.xlsx"

OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
TEMPLATES_DIR.mkdir(exist_ok=True, parents=True)
DATA_DIR.mkdir(exist_ok=True, parents=True)

# Если файла компаний нет — создадим шаблон
if not COMPANIES_XLSX.exists():
    pd.DataFrame(columns=["Компания", "ИНН", "Номера", "Оплата"]).to_excel(COMPANIES_XLSX, index=False)

# === Ячейки шаблона ===
CELL_CUSTOMER = "I5"
CELL_PLATE = "G6"
CELL_DRIVER = "G7"
CELL_DEFECT_LINE1 = "Y8"
CELL_DEFECT_LINE2 = "A9"
CELL_ISSUED_TO = "N10"
CELL_DATE = "CG4"
CELL_TOTAL_NUM = "BR38"
CELL_TOTAL_TEXT = "A39"
CELL_MECHANIC = "W43"

SERVICES_START_ROW = 13
COL_QTY = "BF"
COL_PRICE = "BR"
COL_COST = "CD"

DEFECTS = [
    "Износ автошины",
    "Повреждение автошины",
    "Деформация (грыжа)",
    "Искажение протектора",
    "Трещина на боковой части шины",
    "Вмятина на протекторе",
    "Расслоение и деформация протектора",
    "Разрыв протектора",
    "Разрыв по боковине",
    "Механический разрез боковины",
    "Установка новых автошин",
    "Сезонная перебортировка колёс",
    "Вулканизация",
    "Накачка шин",
    "Другое (ввести вручную)",
]

SERVICES = [
    "Снятие/установка",
    "Мойка",
    "Разбортовка",
    "Забортовка",
    "Балансировка",
    "Установка камеры",
    "Ремонт камеры",
    "Герметик",
    "Ремонт покрышки пласт. №",
    "Снятие запасного колеса",
    "Вулканизация покрышки",
    "Вентиль грузовой",
    "Вентиль ремонтный",
    "Вентиль легковой",
    "Грибок №",
    "Грузики",
    "Удлинитель",
    "Установка вентиля",
    "Флипер",
    "Утилизация",
    "Камера",
    "Подкачка",
    "Жгут",
    "Разгрузка и погрузка колеса",
    "Срочность",
]

SERVICE_NAME_MAP = {
    "Снятие/установка": "Снятие, установка наружное/внутреннее",
    "Вентиль легковой": "Вентиль легковой (хром/черный)",
}

# Базовая цена по умолчанию (если не найдено в таблице)
FIXED_PRICE = 100

# === Прайс из Excel ===

def load_price_data() -> dict:
    """Загружает прайс из файла price.xlsx в удобный словарь."""
    price_file = DATA_DIR / "price.xlsx"
    data = {"truck": {"diameters": [], "services": {}},
            "car": {"diameters": [], "services": {}}}
    try:
        wb = load_workbook(price_file)
        ws = wb.active
    except Exception:
        return data

    # Диаметры
    data["truck"]["diameters"] = [c.value for c in ws[2][1:8] if c.value]
    data["car"]["diameters"] = [c.value for c in ws[2][9:] if c.value]

    row = 3
    while True:
        name = ws.cell(row=row, column=1).value
        if not name:
            break
        truck_prices = {}
        for col, diam in enumerate(data["truck"]["diameters"], start=2):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and "/" in val:
                parts = [p.strip() for p in val.split("/")]
                if len(parts) == 2:
                    try:
                        truck_prices[diam] = (int(parts[0]), int(parts[1]))
                    except ValueError:
                        truck_prices[diam] = val
            else:
                truck_prices[diam] = val
        car_prices = {}
        for col, diam in enumerate(data["car"]["diameters"], start=10):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and "/" in val:
                parts = [p.strip() for p in val.split("/")]
                if len(parts) == 2:
                    try:
                        car_prices[diam] = (int(parts[0]), int(parts[1]))
                    except ValueError:
                        car_prices[diam] = val
            else:
                car_prices[diam] = val
        data["truck"]["services"][name] = truck_prices
        data["car"]["services"][name] = car_prices
        row += 1
    return data

PRICE_DATA = load_price_data()

# === Работа с компаниями ===
COL_NAME = "Компания"
COL_INN = "ИНН"
COL_PLATES = "Номера"
COL_PAY = "Оплата"

def _normalize_company_df(df: pd.DataFrame) -> pd.DataFrame:
    # Поддержка разных заголовков (включая варианты вроде "Оплата (да/нет)")
    mapping = {}
    for col in df.columns:
        v = str(col).strip().lower()
        if v in ("компания", "название", "организация", "контрагент", "company", "name"):
            mapping[col] = COL_NAME
        elif v in ("инн", "inn"):
            mapping[col] = COL_INN
        elif v in ("номера", "госномер", "госномера", "машины", "авто", "plates", "cars"):
            mapping[col] = COL_PLATES
        elif ("оплат" in v) or v in ("оплата", "опл", "pay", "payment"):
            mapping[col] = COL_PAY
    df2 = df.rename(columns=mapping).copy()
    for c in (COL_NAME, COL_INN, COL_PLATES, COL_PAY):
        if c not in df2.columns:
            df2[c] = ""
    df2 = df2[[COL_NAME, COL_INN, COL_PLATES, COL_PAY]]
    for c in (COL_NAME, COL_INN, COL_PLATES, COL_PAY):
        df2[c] = df2[c].astype(str).fillna("").str.strip()
    return df2

def read_companies_df() -> pd.DataFrame:
    try:
        df = pd.read_excel(COMPANIES_XLSX, dtype=str)
    except Exception:
        df = pd.DataFrame(columns=[COL_NAME, COL_INN, COL_PLATES, COL_PAY])
    return _normalize_company_df(df)

def write_companies_df(df: pd.DataFrame):
    # Сохраняем как есть, без сортировки — чтобы новые компании были в конце
    df.to_excel(COMPANIES_XLSX, index=False)

def parse_plates(cell_value: str) -> list[str]:
    return [p.strip() for p in str(cell_value).split(",") if p.strip()]

def join_plates(plates: list[str]) -> str:
    return ", ".join(sorted(set([p.strip() for p in plates if p.strip()])))

def load_companies() -> tuple[dict, list[str]]:
    df = read_companies_df()
    companies = {}
    visible_names = []
    for _, row in df.iterrows():  # сохраняем порядок строк
        name = row[COL_NAME]
        inn = row[COL_INN]
        plates = parse_plates(row[COL_PLATES])
        pay = str(row[COL_PAY]).strip().lower()
        if name:
            companies[name] = {"inn": inn, "plates": plates, "pay": pay}
            if pay in ("да","yes","true","1"):
                visible_names.append(name)
    return companies, visible_names

COMPANIES, ALL_COMPANY_NAMES = load_companies()
PLATE_TO_COMPANY = {p: name for name, meta in COMPANIES.items() for p in meta["plates"]}
ALL_PLATES = list(PLATE_TO_COMPANY.keys())

def reload_companies_globals():
    global COMPANIES, ALL_COMPANY_NAMES, PLATE_TO_COMPANY, ALL_PLATES
    COMPANIES, ALL_COMPANY_NAMES = load_companies()
    PLATE_TO_COMPANY = {p: name for name, meta in COMPANIES.items() for p in meta["plates"]}
    ALL_PLATES = list(PLATE_TO_COMPANY.keys())

# === Чек и текст суммы ===
def ruble_suffix(n: int) -> str:
    n_abs = abs(n) % 100
    n1 = n_abs % 10
    if 11 <= n_abs <= 19:
        return "рублей"
    if n1 == 1:
        return "рубль"
    if 2 <= n1 <= 4:
        return "рубля"
    return "рублей"

def make_total_text(total: int) -> str:
    words = num2words(total, lang='ru').capitalize()
    return f"{words} {ruble_suffix(total)}"

# === Экспорт PDF ===
def export_pdf_via_excel(xlsx_path: Path, pdf_path: Path, a5: bool = True, landscape: bool = False) -> bool:
    try:
        import win32com.client as win32
        from win32com.client import constants
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        ws = wb.Worksheets(1)
        if a5:
            ws.PageSetup.PaperSize = constants.xlPaperA5
        ws.PageSetup.Orientation = constants.xlLandscape if landscape else constants.xlPortrait
        xlTypePDF = 0
        wb.ExportAsFixedFormat(xlTypePDF, str(pdf_path.resolve()))
        wb.Close(SaveChanges=False)
        excel.Quit()
        return True
    except Exception:
        return False

def export_pdf_via_libreoffice(xlsx_path: Path, pdf_path: Path) -> bool:
    try:
        outdir = pdf_path.parent
        cmd = ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(outdir), str(xlsx_path.resolve())]
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        produced = outdir / (xlsx_path.stem + ".pdf")
        if produced.exists():
            if produced != pdf_path:
                produced.replace(pdf_path)
            return True
        return False
    except Exception:
        return False

# === Заполнение шаблона ===
def _write_to_excel(ws, data: dict) -> int:
    ws[CELL_CUSTOMER] = data["customer_display"]
    ws[CELL_PLATE] = data["plate"]
    ws[CELL_DRIVER] = data["driver_name"]

    defect_value = data["defect"]
    ws[CELL_DEFECT_LINE1] = "" if defect_value == "Пропустить" else defect_value
    ws[CELL_DEFECT_LINE2] = ""

    ws[CELL_ISSUED_TO] = data["issued_to"]
    ws[CELL_DATE] = datetime.datetime.now().strftime("%d.%m.%Y")
    ws[CELL_MECHANIC] = data["mechanic"]

    total = 0
    for idx, service_name in enumerate(SERVICES):
        row = SERVICES_START_ROW + idx
        info = data["services"].get(service_name)
        qty = info["qty"] if info else 0
        price = info["price"] if info else 0
        cost = qty * price
        ws[f"{COL_QTY}{row}"] = qty if qty else ""
        ws[f"{COL_PRICE}{row}"] = price if qty else ""
        ws[f"{COL_COST}{row}"] = cost if qty else ""
        total += cost

    ws[CELL_TOTAL_NUM] = total
    ws[CELL_TOTAL_TEXT] = make_total_text(total)
    return total

def fill_excel_only(data: dict) -> Path:
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Не найден шаблон: {TEMPLATE_XLSX}")
    dt = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_out = OUTPUT_DIR / f"наряд_{dt}.xlsx"
    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active
    _write_to_excel(ws, data)
    wb.save(xlsx_out)
    return xlsx_out

def fill_excel_and_export_pdf(data: dict) -> tuple[Path, Path]:
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Не найден шаблон: {TEMPLATE_XLSX}")
    dt = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_out = OUTPUT_DIR / f"наряд_{dt}.xlsx"
    pdf_out = OUTPUT_DIR / f"наряд_{dt}.pdf"
    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active
    _write_to_excel(ws, data)
    wb.save(xlsx_out)
    ok = export_pdf_via_excel(xlsx_out, pdf_out, a5=True, landscape=False)
    if not ok and not export_pdf_via_libreoffice(xlsx_out, pdf_out):
        raise RuntimeError("Не удалось экспортировать в PDF. Проверьте наличие Microsoft Excel (или LibreOffice в PATH).")
    return xlsx_out, pdf_out

# === Скролл-фреймы ===
class VScrollFrame(ttk.Frame):
    def __init__(self, master, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")

        self.inner = ttk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0,0), window=self.inner, anchor="nw")

        self._need_scroll = False

        def _update_scrollregion(event=None):
            self.canvas.itemconfig(self.inner_id, width=self.canvas.winfo_width())
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            need = (self.inner.winfo_reqheight() > self.canvas.winfo_height())
            if need != self._need_scroll:
                self._need_scroll = need
                if self._need_scroll:
                    self.vsb.grid()
                else:
                    self.vsb.grid_remove()
                    self.canvas.yview_moveto(0)

        self.inner.bind("<Configure>", _update_scrollregion)
        self.canvas.bind("<Configure>", _update_scrollregion)

        # колёсико по наведению
        def _bind_wheel(_=None):
            if self._need_scroll:
                self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        def _unbind_wheel(_=None):
            self.canvas.unbind_all("<MouseWheel>")
        for w in (self.canvas, self.inner):
            w.bind("<Enter>", _bind_wheel)
            w.bind("<Leave>", _unbind_wheel)

    def _on_mousewheel(self, event):
        if not self._need_scroll:
            return
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

class HighlightList(tb.Frame):
    def __init__(self, master, on_select, keybind_parent=None):
        super().__init__(master)
        self.on_select = on_select
        self.items = []
        self.current_index = 0
        self.visible = True
        self.keybind_parent = keybind_parent or master
        self._bind_ids = []

        self.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(self, highlightthickness=0, height=160)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")

        self.inner = tb.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self._need_scroll = False
        def _update(event=None):
            self.canvas.itemconfig(self.inner_id, width=self.canvas.winfo_width())
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            need = (self.inner.winfo_reqheight() > self.canvas.winfo_height())
            if need != self._need_scroll:
                self._need_scroll = need
                if need: self.vsb.grid()
                else:
                    self.vsb.grid_remove()
                    self.canvas.yview_moveto(0)
        self.inner.bind("<Configure>", _update)
        self.canvas.bind("<Configure>", _update)

        # колесо по наведению
        def _bind_wheel(_=None):
            if self._need_scroll:
                self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        def _unbind_wheel(_=None):
            self.canvas.unbind_all("<MouseWheel>")
        for w in (self.canvas, self.inner):
            w.bind("<Enter>", _bind_wheel)
            w.bind("<Leave>", _unbind_wheel)

        self._bind_ids.append(self.keybind_parent.bind("<Up>", self._move_up))
        self._bind_ids.append(self.keybind_parent.bind("<Down>", self._move_down))
        self._bind_ids.append(self.keybind_parent.bind("<Return>", self._enter))

    def destroy(self):
        for bid in self._bind_ids:
            try:
                self.keybind_parent.unbind("<Up>", bid)
                self.keybind_parent.unbind("<Down>", bid)
                self.keybind_parent.unbind("<Return>", bid)
            except Exception:
                pass
        super().destroy()

    def show(self):
        self.grid()
        self.visible = True

    def hide(self):
        self.grid_remove()
        self.visible = False

    def set_items(self, names, query):
        for _, row in self.items:
            row.destroy()
        self.items.clear()

        q = (query or "").lower().strip()

        def highlight_text(name: str):
            if not q:
                return name, None, None
            i = name.lower().find(q)
            if i >= 0:
                return name, i, len(q)
            return name, None, None

        for idx, name in enumerate(names):
            text, start, ln = highlight_text(name)
            row = tb.Frame(self.inner)
            row.pack(fill=X, padx=4, pady=2)

            pre = text[:start] if start is not None else text
            match = text[start:start+ln] if start is not None else ""
            post = text[start+ln:] if start is not None else ""

            tb.Label(row, text=pre, anchor="w").pack(side=LEFT)
            if match:
                tb.Label(row, text=match, bootstyle="warning").pack(side=LEFT)
            if post:
                tb.Label(row, text=post, anchor="w").pack(side=LEFT)

            def _click_factory(n=name):
                return lambda e: self.on_select(n)
            row.bind("<Button-1>", _click_factory())
            for child in row.winfo_children():
                child.bind("<Button-1>", _click_factory())

            self.items.append((name, row))

        self.current_index = 0
        self._refresh_active_row()

        if names:
            self.show()
        else:
            self.hide()

    def _refresh_active_row(self):
        for i, (_, row) in enumerate(self.items):
            row.configure(bootstyle=("info" if i == self.current_index else "secondary"))

    def _move_up(self, event=None):
        if not self.visible or not self.items: return
        self.current_index = (self.current_index - 1) % len(self.items)
        self._refresh_active_row()

    def _move_down(self, event=None):
        if not self.visible or not self.items: return
        self.current_index = (self.current_index + 1) % len(self.items)
        self._refresh_active_row()

    def _enter(self, event=None):
        if not self.visible or not self.items: return
        name, _ = self.items[self.current_index]
        self.on_select(name)

# === Приложение ===
class WorkOrderApp:
    def __init__(self, root: tb.Window):
        self.root = root
        self.root.title("Наряд-Заказ — v2.3")
        self.root.geometry("1280x840")

        # Верхняя панель
        topbar = tb.Frame(self.root, padding=8)
        tb.Label(topbar, text="Наряд‑Заказ", font=("-size", 16, "-weight", "bold")).pack(side=LEFT)
        tb.Button(topbar, text="Создать наряд", bootstyle="primary", command=self.open_create_form).pack(side=RIGHT, padx=6)
        tb.Button(topbar, text="Админ‑панель", bootstyle="secondary", command=self.open_admin_panel).pack(side=RIGHT, padx=6)
        tb.Button(topbar, text="Обновить списки", bootstyle="warning", command=self.refresh_lists).pack(side=RIGHT, padx=6)
        topbar.pack(fill=X)

        self.root.bind("<Control-n>", lambda e: self.open_create_form())

        # Плейсхолдер
        self.placeholder = tb.Frame(self.root, padding=20)
        tb.Label(self.placeholder, text="Нажмите «Создать наряд» или Ctrl+N", bootstyle="secondary").pack()
        self.placeholder.pack(fill=BOTH, expand=True)

        self._create_form_window = None  # ссылка, чтобы обновлять виджеты после админки

    def refresh_lists(self):
        reload_companies_globals()
        # если форма открыта — обновим виджеты (с защитой на уничтоженные)
        self._apply_companies_to_form(self._create_form_window)
        messagebox.showinfo("Готово", "Справочник компаний обновлён из файла.", parent=self.root)

    # ===== Создание наряда =====
    def open_create_form(self):
        win = tb.Toplevel(self.root)
        self._create_form_window = win
        win.title("Создать наряд")
        win.geometry("1400x860")
        win.resizable(True, True)
        try:
            win.state('zoomed')
        except Exception:
            pass

        # хоткеи формы
        win.bind("<Control-s>", lambda e: self._build_xlsx_only())
        win.bind("<Control-p>", lambda e: self._build_and_save())
        win.bind("<Escape>", lambda e: win.destroy())
        self._form_parent = win

        # Две панели
        paned = ttk.PanedWindow(win, orient="horizontal")
        paned.pack(fill=BOTH, expand=True, padx=8, pady=8)

        left_wrap = tb.Frame(paned)
        right_wrap = tb.Frame(paned)
        paned.add(left_wrap, weight=1)
        paned.add(right_wrap, weight=1)

        # Прокручиваемые области (grid)
        left_scroll = VScrollFrame(left_wrap)
        right_scroll = VScrollFrame(right_wrap)
        left_scroll.pack(fill=BOTH, expand=True)
        right_scroll.pack(fill=BOTH, expand=True)

        left = left_scroll.inner
        right = right_scroll.inner

        left.grid_columnconfigure(0, weight=1)
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=1)

        pad = {'padx': 8, 'pady': 6}

        # ===== Левая колонка =====
        frm_customer = tb.Labelframe(left, text="Заказчик", padding=8)
        frm_customer.grid(row=0, column=0, sticky="nwe", **pad)
        frm_customer.grid_columnconfigure(1, weight=1)

        self.customer_type = tk.StringVar(value="Частное лицо")
        tb.Radiobutton(frm_customer, text="Частное лицо", variable=self.customer_type, value="Частное лицо", command=self._on_customer_type_changed).grid(row=0, column=0, sticky=NW, padx=4, pady=4)
        tb.Radiobutton(frm_customer, text="Компания", variable=self.customer_type, value="Компания", command=self._on_customer_type_changed).grid(row=0, column=1, sticky=NW, padx=4, pady=4)

        self.search_mode = tk.StringVar(value="company")
        tb.Radiobutton(frm_customer, text="По компании", variable=self.search_mode, value="company", command=lambda: apply_filter()).grid(row=1, column=0, sticky=NW, padx=4, pady=4)
        tb.Radiobutton(frm_customer, text="По номеру", variable=self.search_mode, value="plate", command=lambda: apply_filter()).grid(row=1, column=1, sticky=NW, padx=4, pady=4)

        tb.Label(frm_customer, text="Поиск (Ctrl+F):").grid(row=2, column=0, sticky=NW, padx=4, pady=4)
        self.company_query = tk.StringVar(value="")
        self.entry_company_query = tb.Entry(frm_customer, textvariable=self.company_query)
        self.entry_company_query.grid(row=2, column=1, sticky="we", padx=4, pady=4)

        def focus_search(event=None):
            self.entry_company_query.focus_set()
            self.entry_company_query.selection_range(0, tk.END)
        win.bind("<Control-f>", focus_search)

        def on_pick_company(name):
            if self.search_mode.get() == "company":
                self.company_selected.set(name)
                self._update_company_meta()
            else:
                self.company_selected.set(PLATE_TO_COMPANY.get(name, ""))
                self._update_company_meta()
                self.plate_var.set(name)
                self.plate_list.set(name)

        self.search_results = HighlightList(frm_customer, on_select=on_pick_company, keybind_parent=win)
        self.search_results.grid(row=3, column=0, columnspan=2, sticky="we", padx=2, pady=(0,6))

        tb.Label(frm_customer, text="Компания:").grid(row=4, column=0, sticky=NW, padx=4, pady=4)
        self.company_selected = tk.StringVar(value=(ALL_COMPANY_NAMES[0] if ALL_COMPANY_NAMES else ""))
        self.cmb_company = tb.Combobox(frm_customer, textvariable=self.company_selected, values=ALL_COMPANY_NAMES, state="readonly")
        self.cmb_company.grid(row=4, column=1, sticky="we", padx=4, pady=4)

        tb.Label(frm_customer, text="ИНН:").grid(row=5, column=0, sticky=NW, padx=4, pady=4)
        self.company_inn_var = tk.StringVar(value="")
        tb.Label(frm_customer, textvariable=self.company_inn_var, bootstyle="secondary").grid(row=5, column=1, sticky="w", padx=4, pady=4)

        def apply_filter(*_):
            q = self.company_query.get().strip().lower()
            if self.search_mode.get() == "company":
                values = [name for name in ALL_COMPANY_NAMES if q in name.lower()] if q else list(ALL_COMPANY_NAMES)
                self.cmb_company["values"] = values
                if values:
                    self.cmb_company.set(values[0])
                else:
                    self.cmb_company.set("")
            else:
                values = [p for p in ALL_PLATES if q in p.lower()] if q else list(ALL_PLATES)
            self.search_results.set_items(values[:50], q)
            self._update_company_meta()

        self._company_query_trace = self.company_query.trace_add("write", apply_filter)
        self.cmb_company.bind("<<ComboboxSelected>>", lambda e: self._update_company_meta())
        apply_filter()

        # Госномер
        frm_plate = tb.Labelframe(left, text="Гос. номер", padding=8)
        frm_plate.grid(row=1, column=0, sticky="we", **pad)
        frm_plate.grid_columnconfigure(0, weight=1)
        frm_plate.grid_columnconfigure(1, weight=1)

        self.plate_var = tk.StringVar()
        self.plate_entry = tb.Entry(frm_plate, textvariable=self.plate_var)
        self.plate_list = tb.Combobox(frm_plate, values=[], state="readonly")

        tb.Label(frm_plate, text="Номер (для частного лица — вручную):").grid(row=0, column=0, sticky=NW, padx=4, pady=4)
        self.plate_entry.grid(row=1, column=0, sticky="we", padx=4, pady=4)
        self.plate_list.grid(row=1, column=1, sticky="we", padx=4, pady=4)

        self.trailer_var = tk.StringVar()
        tb.Label(frm_plate, text="Номер прицепа:").grid(row=2, column=0, sticky=NW, padx=4, pady=4)
        tb.Entry(frm_plate, textvariable=self.trailer_var).grid(row=2, column=1, sticky="we", padx=4, pady=4)

        # Водитель
        frm_driver = tb.Labelframe(left, text="Ф.И.О. водителя", padding=8)
        frm_driver.grid(row=2, column=0, sticky="we", **pad)
        self.driver_name = tk.StringVar()
        e = tb.Entry(frm_driver, textvariable=self.driver_name)
        e.grid(row=0, column=0, sticky="we", padx=4, pady=4)
        frm_driver.grid_columnconfigure(0, weight=1)

        # Дефект
        frm_defect = tb.Labelframe(left, text="Описание заказа и дефекта", padding=8)
        frm_defect.grid(row=3, column=0, sticky="we", **pad)
        frm_defect.grid_columnconfigure(1, weight=1)
        self.defect_choice = tk.StringVar(value=DEFECTS[0])
        tb.Label(frm_defect, text="Из списка:").grid(row=0, column=0, sticky=NW, padx=4, pady=4)
        cmb_def = tb.Combobox(frm_defect, textvariable=self.defect_choice, values=DEFECTS, state="readonly")
        cmb_def.grid(row=0, column=1, sticky="we", padx=4, pady=4)
        tb.Label(frm_defect, text="Или 'Другое':").grid(row=1, column=0, sticky=NW, padx=4, pady=4)
        self.defect_custom = tk.StringVar()
        self.defect_entry = tb.Entry(frm_defect, textvariable=self.defect_custom, state=DISABLED)
        self.defect_entry.grid(row=1, column=1, sticky="we", padx=4, pady=4)

        def on_defect_changed(*_):
            if self.defect_choice.get() == "Другое (ввести вручную)":
                self.defect_entry.configure(state=NORMAL)
                self.defect_entry.focus_set()
            else:
                self.defect_entry.configure(state=DISABLED)
                self.defect_custom.set("")
        cmb_def.bind("<<ComboboxSelected>>", lambda e: on_defect_changed())
        on_defect_changed()

        # Исполнители
        frm_people = tb.Labelframe(left, text="Исполнители", padding=8)
        frm_people.grid(row=4, column=0, sticky="we", **pad)
        frm_people.grid_columnconfigure(1, weight=1)
        self.issued_to = tk.StringVar()
        self.mechanic = tk.StringVar()
        tb.Label(frm_people, text="Наряд выдан (фамилия исполнителя):").grid(row=0, column=0, sticky=NW, padx=4, pady=4)
        tb.Entry(frm_people, textvariable=self.issued_to).grid(row=0, column=1, sticky="we", padx=4, pady=4)
        tb.Label(frm_people, text="Фамилия механика:").grid(row=1, column=0, sticky=NW, padx=4, pady=4)
        tb.Entry(frm_people, textvariable=self.mechanic).grid(row=1, column=1, sticky="we", padx=4, pady=4)

        # ===== Правая колонка =====
        frm_vehicle = tb.Labelframe(right, text="Параметры", padding=8)
        frm_vehicle.grid(row=0, column=0, sticky="we", **pad)
        self.vehicle_type = tk.StringVar(value="car")
        tb.Radiobutton(frm_vehicle, text="Легковой", variable=self.vehicle_type, value="car", command=self._update_diameters).grid(row=0, column=0, padx=4, pady=4, sticky=NW)
        tb.Radiobutton(frm_vehicle, text="Грузовой", variable=self.vehicle_type, value="truck", command=self._update_diameters).grid(row=0, column=1, padx=4, pady=4, sticky=NW)
        self.diameter_var = tk.StringVar()
        self.cmb_diameter = tb.Combobox(frm_vehicle, textvariable=self.diameter_var, state="readonly")
        self.cmb_diameter.grid(row=1, column=0, columnspan=2, sticky="we", padx=4, pady=4)
        self.cmb_diameter.bind("<<ComboboxSelected>>", self._update_service_prices)

        frm_services = tb.Labelframe(right, text="Услуги", padding=8)
        frm_services.grid(row=1, column=0, sticky="nsew", **pad)
        right.grid_rowconfigure(1, weight=1)
        frm_services.grid_columnconfigure(0, weight=1)
        frm_services.grid_rowconfigure(1, weight=1)

        # Шапка
        header = tb.Frame(frm_services)
        header.grid(row=0, column=0, sticky="we")
        header.grid_columnconfigure(0, weight=1)
        tb.Label(header, text="Услуга").grid(row=0, column=0, sticky="w", padx=4, pady=2)
        tb.Label(header, text="Кол-во").grid(row=0, column=1, sticky="w", padx=4, pady=2)
        tb.Label(header, text="Опция").grid(row=0, column=2, sticky="w", padx=4, pady=2)
        tb.Label(header, text="Цена").grid(row=0, column=3, sticky="w", padx=4, pady=2)

        # Прокручиваемый список услуг
        svc = VScrollFrame(frm_services)
        svc.grid(row=1, column=0, sticky="nsew", pady=(4,0))
        svc.canvas.configure(height=640)
        svc_inner = svc.inner

        self.services_vars = {}
        self.services_qty = {}
        self.services_price_labels = {}
        self.service_option_vars = {}
        for i, name in enumerate(SERVICES, start=1):
            var = tk.IntVar(value=0)
            qty = tk.IntVar(value=0)
            def _on_toggle_factory(v=var, q=qty):
                def handler():
                    if v.get() and q.get() == 0:
                        q.set(1)
                    if not v.get():
                        q.set(0)
                    self._update_service_prices()
                return handler
            tb.Checkbutton(svc_inner, text=name, variable=var, command=_on_toggle_factory()).grid(row=i, column=0, sticky=NW, padx=4, pady=2)
            tb.Spinbox(svc_inner, from_=0, to=999, textvariable=qty, width=6, command=self._update_service_prices).grid(row=i, column=1, sticky=NW, padx=4, pady=2)
            opt_var = None
            if name == "Снятие/установка":
                opt_var = tk.StringVar(value="наружное")
                cb = tb.Combobox(svc_inner, textvariable=opt_var, values=["наружное","внутреннее"], state="readonly", width=12)
                cb.grid(row=i, column=2, sticky=NW, padx=4, pady=2)
                cb.bind("<<ComboboxSelected>>", self._update_service_prices)
            elif name == "Вентиль легковой":
                opt_var = tk.StringVar(value="чёрный")
                cb = tb.Combobox(svc_inner, textvariable=opt_var, values=["чёрный","хром"], state="readonly", width=12)
                cb.grid(row=i, column=2, sticky=NW, padx=4, pady=2)
                cb.bind("<<ComboboxSelected>>", self._update_service_prices)
            else:
                tb.Label(svc_inner, text="").grid(row=i, column=2, sticky=NW, padx=4, pady=2)
            price_lbl = tb.Label(svc_inner, text="")
            price_lbl.grid(row=i, column=3, sticky=NW, padx=4, pady=2)
            svc_inner.grid_columnconfigure(0, weight=1)
            self.services_vars[name] = var
            self.services_qty[name] = qty
            self.services_price_labels[name] = price_lbl
            if opt_var:
                self.service_option_vars[name] = opt_var

        # Кнопки действия (внизу правой панели)
        actions = tb.Frame(right)
        actions.grid(row=2, column=0, sticky="we", **pad)
        tb.Button(actions, text="Сформировать Excel (Ctrl+S)", bootstyle="success", command=self._build_xlsx_only).pack(side=LEFT, padx=6)
        tb.Button(actions, text="Сформировать PDF (Ctrl+P)", bootstyle="info", command=self._build_and_save).pack(side=LEFT, padx=6)

        # Инициализация
        self._on_customer_type_changed()
        self._update_company_meta()
        self._update_diameters()

        # Корректное отключение trace/биндов при закрытии окна
        def _cleanup():
            try:
                self.company_query.trace_remove("write", self._company_query_trace)
            except Exception:
                pass
            try:
                self.search_results.destroy()
            except Exception:
                pass
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _cleanup)

    # Применить текущий справочник к открытой форме
    def _apply_companies_to_form(self, win):
        # форма может быть не открытой или уже закрыта
        if not hasattr(self, "cmb_company") or not self._widget_exists(self.cmb_company):
            return
        self.cmb_company["values"] = ALL_COMPANY_NAMES
        if ALL_COMPANY_NAMES:
            self.cmb_company.set(ALL_COMPANY_NAMES[0])
        else:
            self.cmb_company.set("")
        # перезаполнить поиск (если виджеты живы)
        if hasattr(self, "company_query"):
            q = self.company_query.get().strip().lower()
            values = [name for name in ALL_COMPANY_NAMES if q in name.lower()] if q else list(ALL_COMPANY_NAMES)
            if hasattr(self, "search_results") and self._widget_exists(self.search_results):
                self.search_results.set_items(values[:50], q)
        self._update_company_meta()

    # ======= Админ‑панель =======
    def open_admin_panel(self):
        # пароль
        pwd = simpledialog.askstring("Вход в админ‑панель", "Введите пароль:", show='*', parent=self.root)
        if pwd != "12345":
            messagebox.showerror("Доступ запрещён", "Неверный пароль.", parent=self.root)
            return

        win = tb.Toplevel(self.root)
        win.title("Админ‑панель")
        win.geometry("1000x700")
        nb = ttk.Notebook(win)
        nb.pack(fill=BOTH, expand=True, padx=8, pady=8)

        # ====== вкладка Добавить компанию ======
        tab_add_company = tb.Frame(nb, padding=10)
        nb.add(tab_add_company, text="Добавить компанию")

        name_var = tk.StringVar()
        inn_var = tk.StringVar()
        plates_var = tk.StringVar()
        tb.Label(tab_add_company, text="Название компании:").grid(row=0, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_company, textvariable=name_var).grid(row=0, column=1, sticky="we", pady=4)
        tb.Label(tab_add_company, text="ИНН:").grid(row=1, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_company, textvariable=inn_var).grid(row=1, column=1, sticky="we", pady=4)
        tb.Label(tab_add_company, text="Гос. номера (через запятую):").grid(row=2, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_company, textvariable=plates_var).grid(row=2, column=1, sticky="we", pady=4)
        tab_add_company.grid_columnconfigure(1, weight=1)

        def do_add_company():
            name = name_var.get().strip()
            inn = inn_var.get().strip()
            plates = join_plates(parse_plates(plates_var.get()))
            if not name:
                messagebox.showerror("Ошибка", "Введите название компании.", parent=win); return
            df = read_companies_df()
            if (df[COL_NAME].str.lower() == name.lower()).any():
                messagebox.showerror("Ошибка", "Компания с таким названием уже существует.", parent=win); return
            # добавляем В КОНЕЦ
            df.loc[len(df)] = {COL_NAME: name, COL_INN: inn, COL_PLATES: plates, COL_PAY: "да"}
            write_companies_df(df)
            reload_companies_globals()
            # обновим GUI, если окно формы открыто
            self._apply_companies_to_form(self._create_form_window)
            # обновим списки во всех вкладках админки
            _apply_filter1(); _apply_filter2(); _apply_filter3(); _apply_filter4(); _refresh_plates_list(); _sync_pay_toggle()
            messagebox.showinfo("Готово", "Компания добавлена (в конец) и включена в списки (Оплата=да).", parent=win)

        tb.Button(tab_add_company, text="Добавить", bootstyle="success", command=do_add_company).grid(row=3, column=1, sticky="e", pady=8)

        # ====== вкладка Добавить гос.номер ======
        tab_add_plate = tb.Frame(nb, padding=10)
        nb.add(tab_add_plate, text="Добавить гос.номер")

        q1 = tk.StringVar()
        tb.Label(tab_add_plate, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q1 = tb.Entry(tab_add_plate, textvariable=q1); e_q1.grid(row=0, column=1, sticky="we", pady=4)
        tab_add_plate.grid_columnconfigure(1, weight=1)
        combo1 = tb.Combobox(tab_add_plate, values=list(COMPANIES.keys()), state="readonly")
        combo1.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        def _apply_filter1(*_):
            all_names = list(COMPANIES.keys())
            qq = q1.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo1["values"] = vals
            if vals:
                combo1.set(vals[0])
        q1.trace_add("write", _apply_filter1)
        _apply_filter1()

        newplates_var = tk.StringVar()
        tb.Label(tab_add_plate, text="Новые номера (через запятую):").grid(row=2, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_plate, textvariable=newplates_var).grid(row=2, column=1, sticky="we", pady=4)

        def do_add_plates():
            name = combo1.get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=win); return
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена в таблице.", parent=win); return
            plates_old = parse_plates(df.loc[mask, COL_PLATES].iloc[0])
            plates_new = parse_plates(newplates_var.get())
            plates_joined = join_plates(plates_old + plates_new)
            df.loc[mask, COL_PLATES] = plates_joined
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter1(); _refresh_plates_list()
            messagebox.showinfo("Готово", "Номера добавлены.", parent=win)

        tb.Button(tab_add_plate, text="Добавить номера", bootstyle="success", command=do_add_plates).grid(row=3, column=1, sticky="e", pady=8)

        # ====== вкладка Оплата on/off ======
        tab_pay = tb.Frame(nb, padding=10)
        nb.add(tab_pay, text="Выставить оплату")

        q2 = tk.StringVar()
        tb.Label(tab_pay, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q2 = tb.Entry(tab_pay, textvariable=q2); e_q2.grid(row=0, column=1, sticky="we", pady=4)
        tab_pay.grid_columnconfigure(1, weight=1)
        combo2 = tb.Combobox(tab_pay, values=list(COMPANIES.keys()), state="readonly")
        combo2.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        pay_var = tk.BooleanVar(value=False)
        tb.Checkbutton(tab_pay, text="Оплата включена (да)", variable=pay_var, bootstyle="success-square-toggle").grid(row=2, column=0, sticky=NW, pady=4)

        def _sync_pay_toggle(*_):
            name = combo2.get().strip()
            if not name:
                pay_var.set(False); return
            df_state = read_companies_df()
            mask = df_state[COL_NAME].str.lower() == name.lower()
            current = str(df_state.loc[mask, COL_PAY].iloc[0]).strip().lower() if mask.any() else ''
            pay_var.set(current in ("да","yes","true","1"))

        def _apply_filter2(*_):
            all_names = list(COMPANIES.keys())
            qq = q2.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo2["values"] = vals
            if vals:
                combo2.set(vals[0])
                _sync_pay_toggle()
        q2.trace_add("write", _apply_filter2); _apply_filter2()
        combo2.bind("<<ComboboxSelected>>", _sync_pay_toggle)

        def do_set_pay():
            name = combo2.get().strip()
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена.", parent=win); return
            df.loc[mask, COL_PAY] = "да" if pay_var.get() else "нет"
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter2(); _sync_pay_toggle()
            messagebox.showinfo("Готово", "Статус оплаты обновлён.", parent=win)

        tb.Button(tab_pay, text="Сохранить", bootstyle="success", command=do_set_pay).grid(row=3, column=1, sticky="e", pady=8)

        # ====== вкладка Удалить компанию ======
        tab_del_company = tb.Frame(nb, padding=10)
        nb.add(tab_del_company, text="Удалить компанию")

        q3 = tk.StringVar()
        tb.Label(tab_del_company, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q3 = tb.Entry(tab_del_company, textvariable=q3); e_q3.grid(row=0, column=1, sticky="we", pady=4)
        tab_del_company.grid_columnconfigure(1, weight=1)
        combo3 = tb.Combobox(tab_del_company, values=list(COMPANIES.keys()), state="readonly")
        combo3.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        def _apply_filter3(*_):
            all_names = list(COMPANIES.keys())
            qq = q3.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo3["values"] = vals
            if vals:
                combo3.set(vals[0])
        q3.trace_add("write", _apply_filter3); _apply_filter3()

        def do_del_company():
            name = combo3.get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=win); return
            if not messagebox.askyesno("Подтвердите", f"Удалить компанию «{name}» и все её номера?", parent=win):
                return
            df = read_companies_df()
            df = df[~(df[COL_NAME].str.lower() == name.lower())]
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter1(); _apply_filter2(); _apply_filter3(); _apply_filter4(); _refresh_plates_list(); _sync_pay_toggle()
            messagebox.showinfo("Готово", "Компания удалена.", parent=win)

        tb.Button(tab_del_company, text="Удалить", bootstyle="danger", command=do_del_company).grid(row=2, column=1, sticky="e", pady=8)

        # ====== вкладка Удалить гос.номер ======
        tab_del_plate = tb.Frame(nb, padding=10)
        nb.add(tab_del_plate, text="Удалить гос. номер")

        q4 = tk.StringVar()
        tb.Label(tab_del_plate, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q4 = tb.Entry(tab_del_plate, textvariable=q4); e_q4.grid(row=0, column=1, sticky="we", pady=4)
        tab_del_plate.grid_columnconfigure(1, weight=1)
        combo4 = tb.Combobox(tab_del_plate, values=list(COMPANIES.keys()), state="readonly")
        combo4.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        listbox = tk.Listbox(tab_del_plate, selectmode="extended", height=12)
        listbox.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=6)
        tab_del_plate.grid_rowconfigure(2, weight=1)

        def _refresh_plates_list(*_):
            name = combo4.get().strip()
            listbox.delete(0, tk.END)
            if name and name in COMPANIES:
                for p in COMPANIES[name]["plates"]:
                    listbox.insert(tk.END, p)

        def _apply_filter4(*_):
            all_names = list(COMPANIES.keys())
            qq = q4.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo4["values"] = vals
            if vals:
                combo4.set(vals[0])
                _refresh_plates_list()

        q4.trace_add("write", _apply_filter4); _apply_filter4()
        combo4.bind("<<ComboboxSelected>>", lambda e: _refresh_plates_list())

        def do_del_plates():
            name = combo4.get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=win); return
            sel = [listbox.get(i) for i in listbox.curselection()]
            if not sel:
                messagebox.showerror("Ошибка", "Выберите номера для удаления.", parent=win); return
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена в таблице.", parent=win); return
            old = parse_plates(df.loc[mask, COL_PLATES].iloc[0])
            new = [p for p in old if p not in sel]
            df.loc[mask, COL_PLATES] = join_plates(new)
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter4(); _refresh_plates_list()
            messagebox.showinfo("Готово", "Выбранные номера удалены.", parent=win)

        tb.Button(tab_del_plate, text="Удалить отмеченные номера", bootstyle="danger", command=do_del_plates).grid(row=3, column=1, sticky="e", pady=8)

    # ======= ЛОГИКА формы =======
    def _widget_exists(self, w) -> bool:
        try:
            return bool(w and w.winfo_exists())
        except Exception:
            return False

    def _update_company_meta(self):
        name = getattr(self, "company_selected", tk.StringVar()).get()
        meta = COMPANIES.get(name, {"inn": "", "plates": []})
        if hasattr(self, "company_inn_var"):
            self.company_inn_var.set(meta.get("inn", ""))
        if hasattr(self, "plate_list") and self._widget_exists(self.plate_list):
            plates = meta.get("plates", [])
            self.plate_list["values"] = plates
            if plates:
                self.plate_list.set(plates[0])
            else:
                self.plate_list.set("")

    def _on_customer_type_changed(self):
        is_company = (self.customer_type.get() == "Компания")
        if hasattr(self, "plate_entry") and hasattr(self, "plate_list"):
            if is_company:
                self.plate_entry.configure(state=DISABLED)
                self.plate_list.configure(state="readonly")
            else:
                self.plate_entry.configure(state=NORMAL)
                self.plate_list.configure(state=DISABLED)

    def _update_diameters(self):
        vt = "truck" if self.vehicle_type.get() == "truck" else "car"
        values = PRICE_DATA.get(vt, {}).get("diameters", [])
        if hasattr(self, "cmb_diameter"):
            self.cmb_diameter["values"] = values
            if values:
                self.cmb_diameter.set(values[0])
            else:
                self.cmb_diameter.set("")
        self._update_service_prices()

    def _get_unit_price(self, service_name: str) -> int:
        vt = "truck" if self.vehicle_type.get() == "truck" else "car"
        diam = getattr(self, "diameter_var", tk.StringVar()).get()
        svc_name = SERVICE_NAME_MAP.get(service_name, service_name)
        prices = PRICE_DATA.get(vt, {}).get("services", {}).get(svc_name, {})
        price = prices.get(diam, 0)
        if isinstance(price, tuple):
            opt_var = self.service_option_vars.get(service_name)
            if opt_var:
                opt = opt_var.get()
                if service_name == "Снятие/установка":
                    price = price[1] if opt == "внутреннее" else price[0]
                elif service_name == "Вентиль легковой":
                    price = price[1] if opt == "хром" else price[0]
        return int(price) if price else 0

    def _update_service_prices(self, *_):
        for name in SERVICES:
            price = self._get_unit_price(name)
            lbl = self.services_price_labels.get(name)
            if lbl:
                lbl.configure(text=str(price) if price else "")

    def _collect_services(self) -> dict[str, dict]:
        selected = {}
        for name in SERVICES:
            var = self.services_vars[name]
            qty = max(0, int(self.services_qty[name].get()))
            if var.get() and qty > 0:
                selected[name] = {"qty": qty, "price": self._get_unit_price(name)}
        return selected

    def _validate(self) -> tuple[bool, str]:
        if self.customer_type.get() == "Компания":
            if not getattr(self, "company_selected", tk.StringVar()).get():
                return False, "Выберите компанию."
            if self.company_selected.get() not in ALL_COMPANY_NAMES:
                return False, "Компания недоступна (возможно, Оплата=нет)."
            if not self.plate_list.get():
                return False, "Выберите гос. номер из списка."
        else:
            if not self.plate_entry.get().strip():
                return False, "Введите гос. номер для частного лица."

        if not self.driver_name.get().strip():
            return False, "Введите Ф.И.О. водителя."

        if self.defect_choice.get() == "Другое (ввести вручную)":
            if not self.defect_custom.get().strip():
                return False, "Введите текст дефекта в поле 'Другое'."

        if not self.issued_to.get().strip():
            return False, "Введите фамилию исполнителя ('Наряд выдан')."
        if not self.mechanic.get().strip():
            return False, "Введите фамилию механика."
        if len(self._collect_services()) == 0:
            return False, "Выберите хотя бы одну услугу и укажите количество."
        return True, ""

    def _gather_data(self) -> dict:
        is_company = (self.customer_type.get() == "Компания")
        if is_company:
            customer_display = self.company_selected.get()
            plate_value = self.plate_list.get().strip()
        else:
            customer_display = "Частное лицо"
            plate_value = self.plate_entry.get().strip()
        trailer = self.trailer_var.get().strip()
        if trailer:
            plate_value = f"{plate_value}, {trailer}" if plate_value else trailer

        if self.defect_choice.get() == "Другое (ввести вручную)":
            defect_value = self.defect_custom.get().strip()
        else:
            defect_value = self.defect_choice.get()

        data = {
            "customer_display": customer_display,
            "plate": plate_value,
            "driver_name": self.driver_name.get().strip(),
            "defect": defect_value,
            "issued_to": self.issued_to.get().strip(),
            "mechanic": self.mechanic.get().strip(),
            "services": self._collect_services(),
            "vehicle_type": self.vehicle_type.get(),
            "diameter": self.diameter_var.get(),
        }
        return data

    def _build_xlsx_only(self):
        ok, msg = self._validate()
        if not ok:
            messagebox.showerror("Ошибка", msg, parent=self._form_parent)
            return
        data = self._gather_data()
        try:
            xlsx_path = fill_excel_only(data)
            messagebox.showinfo("Готово", f"Excel сформирован:\n\n{xlsx_path}\n\nОткрываю папку с результатами.", parent=self._form_parent)
            try:
                os.startfile(str(OUTPUT_DIR.resolve()))
            except Exception:
                pass
        except FileNotFoundError as e:
            messagebox.showerror("Шаблон не найден", str(e), parent=self._form_parent)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Неожиданная ошибка: {e}", parent=self._form_parent)

    def _build_and_save(self):
        ok, msg = self._validate()
        if not ok:
            messagebox.showerror("Ошибка", msg, parent=self._form_parent)
            return
        data = self._gather_data()
        try:
            xlsx_path, pdf_path = fill_excel_and_export_pdf(data)
            messagebox.showinfo("Готово", f"Файлы сохранены:\n\n{xlsx_path}\n{pdf_path}\n\nОткрываю папку с результатами.", parent=self._form_parent)
            try:
                os.startfile(str(OUTPUT_DIR.resolve()))
            except Exception:
                pass
        except FileNotFoundError as e:
            messagebox.showerror("Шаблон не найден", str(e), parent=self._form_parent)
        except RuntimeError as e:
            messagebox.showerror("Не удалось создать PDF", f"{e}\nПроверьте наличие Microsoft Excel (или LibreOffice).", parent=self._form_parent)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Неожиданная ошибка: {e}", parent=self._form_parent)

def main():
    app = tb.Window(themename="flatly")
    WorkOrderApp(app)
    app.mainloop()

if __name__ == "__main__":
    main()
